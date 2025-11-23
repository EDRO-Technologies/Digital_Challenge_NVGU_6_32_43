import os
import logging
from openpyxl import load_workbook
import xlrd
from database.db import add_schedule, add_specialty
from config import EXCEL_FOLDER_1, EXCEL_FOLDER_2

logger = logging.getLogger(__name__)


async def parse_excel_file(file_path: str, specialty_name: str):
    """Парсинг Excel файла и добавление данных в БД"""
    try:
        # Определяем расширение файла
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext == '.xlsx':
            # Используем openpyxl для .xlsx
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Пытаемся найти заголовки
            headers = []
            header_row = None
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_values = [str(cell).strip() if cell else '' for cell in row]
                # Ищем строку с заголовками (обычно содержит "День", "Время", "Предмет" и т.д.)
                if any(keyword in ' '.join(row_values).lower() for keyword in ['день', 'время', 'предмет', 'преподаватель', 'аудитория']):
                    headers = row_values
                    header_row = row_idx
                    break
            
            if not headers:
                # Если заголовки не найдены, используем первую строку
                headers = [str(cell).strip() if cell else '' for cell in next(ws.iter_rows(values_only=True))]
                header_row = 1
            
            # Определяем индексы колонок
            day_col = None
            time_col = None
            subject_col = None
            teacher_col = None
            room_col = None
            group_col = None
            
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if 'день' in header_lower or 'день недели' in header_lower:
                    day_col = idx
                elif 'время' in header_lower or 'час' in header_lower:
                    time_col = idx
                elif 'предмет' in header_lower or 'дисциплина' in header_lower:
                    subject_col = idx
                elif 'преподаватель' in header_lower or 'преп' in header_lower:
                    teacher_col = idx
                elif 'аудитория' in header_lower or 'кабинет' in header_lower or 'комната' in header_lower:
                    room_col = idx
                elif 'группа' in header_lower:
                    group_col = idx
            
            # Парсим данные
            added_count = 0
            for row_idx, row in enumerate(ws.iter_rows(values_only=True, min_row=header_row + 1), header_row + 1):
                row_values = [str(cell).strip() if cell else '' for cell in row]
                
                # Пропускаем пустые строки
                if not any(row_values):
                    continue
                
                day = row_values[day_col] if day_col is not None and day_col < len(row_values) else None
                time = row_values[time_col] if time_col is not None and time_col < len(row_values) else None
                subject = row_values[subject_col] if subject_col is not None and subject_col < len(row_values) else None
                teacher = row_values[teacher_col] if teacher_col is not None and teacher_col < len(row_values) else None
                room = row_values[room_col] if room_col is not None and room_col < len(row_values) else None
                group = row_values[group_col] if group_col is not None and group_col < len(row_values) else None
                
                if day and time and subject:
                    await add_schedule(
                        specialty=specialty_name,
                        day_of_week=day,
                        time=time,
                        subject=subject,
                        teacher=teacher or None,
                        room=room or None,
                        group_name=group or None
                    )
                    added_count += 1
            
            return added_count
            
        elif ext == '.xls':
            # Используем xlrd для .xls
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)
            
            # Ищем заголовки
            headers = []
            header_row = 0
            
            for row_idx in range(sheet.nrows):
                row_values = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
                if any(keyword in ' '.join(row_values).lower() for keyword in ['день', 'время', 'предмет', 'преподаватель']):
                    headers = row_values
                    header_row = row_idx
                    break
            
            if not headers:
                headers = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
            
            # Определяем индексы колонок
            day_col = None
            time_col = None
            subject_col = None
            teacher_col = None
            room_col = None
            group_col = None
            
            for idx, header in enumerate(headers):
                header_lower = header.lower()
                if 'день' in header_lower:
                    day_col = idx
                elif 'время' in header_lower:
                    time_col = idx
                elif 'предмет' in header_lower:
                    subject_col = idx
                elif 'преподаватель' in header_lower:
                    teacher_col = idx
                elif 'аудитория' in header_lower:
                    room_col = idx
                elif 'группа' in header_lower:
                    group_col = idx
            
            # Парсим данные
            added_count = 0
            for row_idx in range(header_row + 1, sheet.nrows):
                row_values = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
                
                if not any(row_values):
                    continue
                
                day = row_values[day_col] if day_col is not None and day_col < len(row_values) else None
                time = row_values[time_col] if time_col is not None and time_col < len(row_values) else None
                subject = row_values[subject_col] if subject_col is not None and subject_col < len(row_values) else None
                teacher = row_values[teacher_col] if teacher_col is not None and teacher_col < len(row_values) else None
                room = row_values[room_col] if room_col is not None and room_col < len(row_values) else None
                group = row_values[group_col] if group_col is not None and group_col < len(row_values) else None
                
                if day and time and subject:
                    await add_schedule(
                        specialty=specialty_name,
                        day_of_week=day,
                        time=time,
                        subject=subject,
                        teacher=teacher or None,
                        room=room or None,
                        group_name=group or None
                    )
                    added_count += 1
            
            return added_count
            
    except Exception as e:
        logger.error(f"Ошибка при парсинге файла {file_path}: {str(e)}")
        import traceback
        logger.debug(traceback.format_exc())
        return 0


async def load_all_excel_files():
    """Загрузить все Excel файлы из папок"""
    folders = [EXCEL_FOLDER_1, EXCEL_FOLDER_2]
    total_added = 0
    files_processed = 0
    
    logger.info(f"Начало загрузки Excel файлов из папок: {folders}")
    
    for folder in folders:
        if not os.path.exists(folder):
            logger.warning(f"Папка {folder} не существует")
            continue
        
        logger.info(f"Обработка папки: {folder}")
        files_in_folder = [f for f in os.listdir(folder) if f.endswith(('.xls', '.xlsx'))]
        logger.info(f"Найдено файлов в {folder}: {len(files_in_folder)}")
        
        for filename in files_in_folder:
            try:
                file_path = os.path.join(folder, filename)
                logger.info(f"Обработка файла: {filename}")
                
                # Извлекаем название специальности из имени файла
                specialty_name = os.path.splitext(filename)[0]
                logger.info(f"Специальность: {specialty_name}")
                
                # Добавляем специальность в БД
                await add_specialty(specialty_name)
                
                # Парсим файл
                added = await parse_excel_file(file_path, specialty_name)
                total_added += added
                files_processed += 1
                logger.info(f"Файл {filename} обработан: добавлено {added} записей")
            except Exception as e:
                logger.error(f"Ошибка при обработке файла {filename}: {str(e)}")
                continue
    
    logger.info(f"Загрузка завершена. Обработано файлов: {files_processed}, добавлено записей: {total_added}")
    return total_added

